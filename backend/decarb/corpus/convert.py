"""
Convert raw corpus files (PDF, HTML, XLSX) to markdown.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

log = logging.getLogger("corpus.convert")


def convert_pdf(path: Path) -> list[dict[str, Any]]:
    """Convert PDF to list of {text, page_number} dicts, one per page."""
    import pymupdf4llm

    pages = pymupdf4llm.to_markdown(str(path), page_chunks=True)
    result = []
    for i, page in enumerate(pages):
        text = page.get("text", "") if isinstance(page, dict) else str(page)
        if text.strip():
            page_num = page.get("metadata", {}).get("page", i + 1) if isinstance(page, dict) else i + 1
            result.append({"text": text, "page_number": page_num})
    return result


def convert_html(path: Path) -> list[dict[str, Any]]:
    """Convert HTML to markdown. Returns single block (no page numbers)."""
    from markdownify import markdownify as md

    html = path.read_text(errors="replace")

    # Strip nav/footer/cookie elements before converting
    from bs4 import BeautifulSoup
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup.find_all(["nav", "footer", "header"]):
        tag.decompose()
    for tag in soup.find_all(attrs={"class": lambda c: c and any(
            x in str(c).lower() for x in ["cookie", "banner", "nav", "footer", "sidebar"])}):
        tag.decompose()

    clean_html = str(soup)
    text = md(clean_html, heading_style="ATX", strip=["img", "script", "style"])

    if text.strip():
        return [{"text": text, "page_number": None}]
    return []


def convert_xlsx(path: Path) -> list[dict[str, Any]]:
    """Convert XLSX sheets to markdown tables. One block per sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"## {sheet_name}\n"]
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            cells = []
            for cell in row:
                val = str(cell) if cell is not None else ""
                if len(val) > 2000:
                    val = val[:2000] + "…"
                cells.append(val.replace("|", "\\|").replace("\n", " "))
            rows_data.append(cells)

        if not rows_data:
            continue

        # Header row
        header = rows_data[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows_data[1:]:
            # Pad/trim to header length
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded[:len(header)]) + " |")

        text = "\n".join(lines)
        if text.strip():
            results.append({"text": text, "page_number": None})

    wb.close()
    return results


def convert_file(path: Path) -> list[dict[str, Any]]:
    """Dispatch to the right converter based on file extension."""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return convert_pdf(path)
    elif ext in (".html", ".htm"):
        return convert_html(path)
    elif ext == ".xlsx":
        return convert_xlsx(path)
    else:
        log.warning(f"Unsupported file type: {ext} — skipping {path.name}")
        return []
